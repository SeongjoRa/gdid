import streamlit as st
from streamlit_option_menu import option_menu
import os, sys
import pandas as pd
import numpy as np
from openpyxl import *
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border,Font, PatternFill, Side
from openpyxl.utils.cell import get_column_letter

st.set_page_config(
    page_title="GPIM Automation Hub",
    layout="centered",
    initial_sidebar_state="auto",        
)

hide_st_style = '''<style>
                    #MainMenu {visibility: hidden;}
                    footer {visibility: hidden;}
                    header {visibility: hidden;}
                    </style>
                    '''
st.markdown(hide_st_style, unsafe_allow_html=True)

if 'greeting' not in st.session_state:
    st.session_state.greeting = "Hello, there!"
# st.write(st.session_state)

with st.sidebar:
    selected = option_menu(menu_title=None, options = ["GDID", "Settings"], 
                        icons=['file-earmark-fill', 'gear'], menu_icon="list", default_index=0) # orientation="horizontal"


wb0 = load_workbook("Settings.xlsx", data_only=True)
ws0 = wb0.active

main_locales = []
for x in range(2, ws0.max_column + 1):
    if not x == "":
        main_locales.append(ws0.cell(row=2, column=x).value)

download_folder = ws0["B1"].value # Default download location

if selected == "Settings":            
    col1, col2, col3 = st.columns(3)
    with col3:
        openxl = st.button("Open Settings file", type="secondary", use_container_width=True)
        if openxl:
            os.startfile("Settings.xlsx") 

    tab1, tab2 = st.tabs(["GDID template Generator","Coming soon"])
    with tab1:
        col1, col2 = st.columns(2)                 
        with col1:
            new_locale = st.text_input("Add any new Locale here.", key="new_locale", placeholder="Add any new Locale here.", label_visibility="collapsed")
        with col2:   
            new_location = st.text_input("Update your download location here.", key="new_location", placeholder="Update your download location here.", label_visibility="collapsed")

        saved = st.button("Save", type="primary", use_container_width=True)
        if saved:
            if not new_locale == "":
                if new_locale not in main_locales:
                    ws0.cell(row=2, column= ws0.max_column + 1).value = new_locale
                    wb0.save("Settings.xlsx")
                    st.info("New settings are saved.", icon="ℹ️")
                else:
                    st.warning("Locale already exists.", icon="ℹ️")

            if not new_location == "": 
                if not new_location == download_folder:
                    ws0["B1"].value = new_location
                    wb0.save("Settings.xlsx")
                    st.info("New settings are saved.", icon="ℹ️")
                else:
                    st.warning("Default download location is the same.", icon="ℹ️")


if selected == "GDID":            
    st.title("GDID template Generator")
    st.write("######")

    locale = st.selectbox('Please select the main Locale.', (main_locales))
    st.write("######")
    uploaded_file = st.file_uploader("Please upload 'GDID' report and click 'Generate' button below.", type=["xlsx"], accept_multiple_files=False, label_visibility="visible")
    submitted = st.button("Generate", type="primary", use_container_width=True)   

    if submitted:
        if uploaded_file is not None:
            filename = str(uploaded_file.name)           
            if not filename.find("GDID") == -1:
                df0 = pd.read_excel( uploaded_file, sheet_name=0, engine='openpyxl')

                rows_cnt = df0.shape[0] # Number of Rows == rows_cnt
                skus = list(df0['SKU'])
                st.write("SKU Count :", rows_cnt)

                if df0['SKU'].isnull().sum() == rows_cnt: 
                    df0.insert(4, '10/11?', np.nan)
                else:
                    df0.insert(4, '10/11?', list(map(lambda x: len(str(x)), skus)))

                headers0 = list(df0.columns.values)
                locale_included = [col for col in headers0 if locale in col]
                if len(locale_included) == 0:
                    st.warning("Report does NOT contain the main Locale.", icon="ℹ️")
                    sys.exit()
 
                headers1 = headers0.copy()
                headers1.remove('Division')
                headers1.remove('Commodity Class')
                headers1.remove('GPH Path')
                headers1.remove('Marketplace Formal Name')

                st.write("Started arranging columns in order!")
                mpfns=[]; dscs=[]; ldscs=[]; b1s=[]; b2s=[]; b3s=[]; b4s=[]; b5s=[]; b6s=[]; b7s=[]; b8s=[]; b9s=[]; b10s=[]; b11s=[]; b12s=[]; b13s=[]; b14s=[]; b15s=[]; pics1=[]; pics2=[]; docs=[]; brnds=[]; mkt1s=[]; mkt2s=[]; mkt3s=[]
                for i in headers0: # Split columns by Attributes(column names) required to update Values for
                    if i.endswith(" — Marketplace Formal Name"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            mpfns.insert(0, i) # Move the column of target locale to the first item in the list
                        else:
                            mpfns.append(i)
                    elif i.endswith(" — Marketplace Description"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            dscs.insert(0, i)
                        else:
                            dscs.append(i)
                    elif i.endswith(" — Marketplace Description Extended"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            ldscs.insert(0, i)
                        else:
                            ldscs.append(i)
                    elif i.endswith(" — Bullet 1"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b1s.insert(0, i)
                        else:
                            b1s.append(i) # List for Bullet 1 columns
                    elif i.endswith(" — Bullet 2"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b2s.insert(0, i)
                        else:
                            b2s.append(i)
                    elif i.endswith(" — Bullet 3"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b3s.insert(0, i)
                        else:
                            b3s.append(i)
                    elif i.endswith(" — Bullet 4"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b4s.insert(0, i)
                        else:
                            b4s.append(i)
                    elif i.endswith(" — Bullet 5"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b5s.insert(0, i)
                        else:
                            b5s.append(i)
                    elif i.endswith(" — Bullet 6"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b6s.insert(0, i)
                        else:
                            b6s.append(i)
                    elif i.endswith(" — Bullet 7"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b7s.insert(0, i)
                        else:
                            b7s.append(i)
                    elif i.endswith(" — Bullet 8"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b8s.insert(0, i)
                        else:
                            b8s.append(i)
                    elif i.endswith(" — Bullet 9"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b9s.insert(0, i)
                        else:
                            b9s.append(i)
                    elif i.endswith(" — Bullet 10"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b10s.insert(0, i)
                        else:
                            b10s.append(i)
                    elif i.endswith(" — Bullet 11"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b11s.insert(0, i)
                        else:
                            b11s.append(i)
                    elif i.endswith(" — Bullet 12"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b12s.insert(0, i)
                        else:
                            b12s.append(i)
                    elif i.endswith(" — Bullet 13"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b13s.insert(0, i)
                        else:
                            b13s.append(i)
                    elif i.endswith(" — Bullet 14"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b14s.insert(0, i)
                        else:
                            b14s.append(i)
                    elif i.endswith(" — Bullet 15"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            b15s.insert(0, i)
                        else:
                            b15s.append(i)
                    elif i.endswith(" — Brand"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            brnds.insert(0, i)
                        else:
                            brnds.append(i)
                    elif i.endswith("FUZE Market Level 1"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            mkt1s.insert(0, i)
                        else:
                            mkt1s.append(i)
                    elif i.endswith("FUZE Market Level 2"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            mkt2s.insert(0, i)
                        else:
                            mkt2s.append(i)
                    elif i.endswith("FUZE Market Level 3"):
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            mkt3s.insert(0, i)
                        else:
                            mkt3s.append(i)
                    elif not i.find(" — Main Picture") == -1:
                        headers1.remove(i)
                        if not i.find(locale) == -1:
                            pics1.insert(0, i)
                        else:
                            pics1.append(i)  

                headers2 = headers1.copy()
                for j in headers1: # Remove columns of Null Values
                    if j in ['10/11?', 'Commodity Class', 'Variant SKU', 'SKU']:
                        continue
                    if df0[j].isnull().sum() == rows_cnt:
                        headers2.remove(j)

                headers3 = headers2.copy()
                for k in headers2:               
                    if not k.find(" — Additional Picture") == -1:
                        headers3.remove(k)
                        pics2.append(k)
                    elif not k.find(" — Doc") == -1:
                        headers3.remove(k)
                        docs.append(k)

                bullets0 = [b5s, b6s, b7s, b8s, b9s, b10s, b11s, b12s, b13s, b14s, b15s]             
                bullets1 = bullets0.copy()
                mkts0 = [mkt2s, mkt3s]
                mkts1 = mkts0.copy()

                for bullet in bullets0: # to check if Vales are available for each Attribute and delte the relevant columns               
                    nan1 = df0[bullet].isnull().sum()
                    if nan1.sum()/len(nan1) == rows_cnt:
                        bullets1.remove(bullet)
                bullets2 = [second for first in bullets1 for second in first]

                for mkt in mkts0:
                    nan3 = df0[mkt].isnull().sum()
                    if nan3.sum()/len(nan3) == rows_cnt:
                        mkts1.remove(mkt)        
                mkts2 = [second for first in mkts1 for second in first]
                st.write("Columns of Null values are deleted.")
            
                headers3[5:5] = docs
                headers3[5:5] = pics2
                headers3[5:5] = brnds
                headers3[5:5] = mkts2
                headers3[5:5] = mkt1s
                headers3[5:5] = pics1
                headers3[5:5] = bullets2
                headers3[5:5] = b4s
                headers3[5:5] = b3s
                headers3[5:5] = b2s
                headers3[5:5] = b1s
                if not df0[ldscs].isnull().sum().sum()/len(df0[ldscs].isnull().sum()) == rows_cnt:
                    headers3[5:5] = ldscs
                headers3[5:5] = dscs
                headers3[5:5] = mpfns
                headers3.insert(1, 'GPH Path')
                headers3.insert(1, 'Commodity Class')
                headers3.insert(1, 'Division')
                st.write("The rest of columns are now reordered.")
        
                df1 = df0[headers3]
                df1 = df1.astype({"Base SKU": 'str'})
                df1.sort_values(by='Base SKU', ascending=True, inplace=True) # Sort by Base coulmn
                df1.to_excel("Digitization Template.xlsx", index=False)

                wb = load_workbook("Digitization Template.xlsx", data_only=True)
                ws = wb.active 

                ws.insert_rows(1) # for validation/gap filling required   
                ws.row_dimensions[2].height = 40
                for r1 in range(3, ws.max_row + 1):
                    ws.row_dimensions[r1].height = 20    

                for c1 in range(1, ws.max_column + 1): # Set styles for header rows (row 2)
                    ws.cell(row=2, column=c1).fill = PatternFill(fgColor="D4D4D4", fill_type="solid")
                    ws.cell(row=2, column=c1).font = Font(name="Arial", size=9, bold=True, color="000000")
                    ws.cell(row=2, column=c1).alignment = Alignment(horizontal="center",wrap_text=True)
                    ws.cell(row=2, column=c1).border = Border(left=Side(style="thin", color="808080"), right=Side(style="thin", color="808080"),top=Side(style="thin", color="808080"),bottom=Side(style="thin", color="808080"))

                ws.column_dimensions["A"].width = 8
                ws.column_dimensions["H"].width = 8 
                for c2 in range(2, 8):
                    ws.column_dimensions[get_column_letter(c2)].width = 12
                    
                for c3 in range(9, ws.max_column + 1):
                    if ws.cell(row=2, column=c3).value == "Locale":
                        break
                    ws.column_dimensions[get_column_letter(c3)].width = 20
                    if not ws.cell(row=2, column=c3).value.find(locale) == -1:
                        ws.cell(row=2, column=c3).fill = PatternFill(fgColor="FFC000", fill_type="solid")
                            
                mpfns_idx=[]; dscs_idx=[]; b1s_idx=[]; b2s_idx=[]; b2s_idx=[]; b3s_idx=[]; b4s_idx=[]; pics1_idx=[]; mkt1s_idx=[]; vals=[]
                for idx, attr in enumerate(ws[2]):
                    if attr.value in mpfns:
                        mpfns_idx.append(idx+1)
                    elif attr.value in dscs:
                        dscs_idx.append(idx+1)
                    elif attr.value in b1s:
                        b1s_idx.append(idx+1)
                    elif attr.value in b2s:
                        b2s_idx.append(idx+1)
                    elif attr.value in b3s:
                        b3s_idx.append(idx+1)
                    elif attr.value in b4s:
                        b4s_idx.append(idx+1)
                    elif attr.value in pics1:
                        pics1_idx.append(idx+1)
                    elif attr.value in mkt1s:
                        mkt1s_idx.append(idx+1)

                def coloring (mindx, maxdx):
                    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=mindx, max_col=maxdx):
                        if not row[0].value is None: 
                            continue
                        vals.clear()            
                        for cell in row: 
                            if not cell.value is None:
                                vals.append(cell.value)
                        if len(vals) < 1:
                            row[0].fill = PatternFill(fgColor="FECCCC", fill_type="solid") # Color the Target Locale Cell RED
                        else:
                            row[0].fill = PatternFill(fgColor="FFF1CD", fill_type="solid") # Color the Target Locale Cell YELLOW        

                coloring (mpfns_idx[0], mpfns_idx[-1])
                coloring (dscs_idx[0], dscs_idx[-1])
                coloring (b1s_idx[0], b1s_idx[-1])
                coloring (b2s_idx[0], b2s_idx[-1])
                coloring (b3s_idx[0], b3s_idx[-1])
                coloring (b4s_idx[0], b4s_idx[-1])
                coloring (pics1_idx[0], pics1_idx[-1])
                coloring (mkt1s_idx[0], mkt1s_idx[-1])

                label_cols = [mpfns_idx[0], dscs_idx[0], b1s_idx[0], b2s_idx[0], b2s_idx[0], b3s_idx[0], b4s_idx[0], pics1_idx[0], mkt1s_idx[0]]
                for c4 in label_cols: # Set styles for columns for Target Locale (Labeling)  
                    ws.cell(row=1, column=c4).fill = PatternFill(fgColor="002060", fill_type="solid")
                    ws.cell(row=1, column=c4).value = 'Validation/Gap Filling Required'
                    ws.cell(row=1, column=c4).font = Font(name= "Arial", size=9, color="FFFFFF")
                    ws.cell(row=1, column=c4).alignment = Alignment(horizontal="center",wrap_text=True)
                st.write("Styles for header rows & color coding for blank cells are applied.")
                
                st.write("Just started applying styles for the data cells.")
                for row2 in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col = ws.max_column):
                    for cell2 in row2:
                        cell2.alignment = Alignment(wrap_text=True)
                        cell2.font = Font(name="Arial", size=9)            
                        cell2.border = Border(left=Side(style="thin", color="808080"), right=Side(style="thin", color="808080"),top=Side(style="thin", color="808080"),bottom=Side(style="thin", color="808080"))        

                ws.freeze_panes = ws["I3"]
                ws.auto_filter.ref = "A2:{}{}".format(get_column_letter(ws.max_column), ws.max_row)
            
                wb.save("Digitization Template.xlsx")
                os.startfile("Digitization Template.xlsx")
            else:
                st.warning("You may have uploded a wrong report.", icon="ℹ️")
        else:
            st.warning("Please upload GDID report first.", icon="ℹ️")

